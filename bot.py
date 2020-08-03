# region imports
import os
import difflib
import re
import datetime
import aiohttp

from discord.ext import commands
import discord

import openpyxl

from bs4 import BeautifulSoup
# endregion

CMD_PREFIX = "!"
HELP_PREFIX = f"``{CMD_PREFIX}"


bot = commands.Bot(command_prefix=CMD_PREFIX)
bot.owner_id = 226700060308668420

# region excel parse
sheet = openpyxl.load_workbook('./islands.xlsx', data_only=True)['Sheet1']
i = 2
islandspos = {}
engislandpos = {}
while (True):
    islandname = sheet.cell(row=i, column=1).value

    if islandname is None:
        break
    pos = sheet.cell(row=i, column=2).value
    region = sheet.cell(row=i, column=3).value
    engname = sheet.cell(row=i, column=4).value

    islandspos[islandname] = [pos, region, engname]
    engislandpos[engname] = [pos, region, engname]
    i += 1
# endregion

# region commands


@bot.command(description="xbox 아이디가 포함되게 닉네임을 변경합니다.",
             help=f"{HELP_PREFIX}아이디`` ``xboxid``")
async def 아이디(ctx, *, id):
    prevnick = ctx.author.display_name

    pattern = r'(?P<nickname>\S+)\s*\(.+\)'

    mat = re.match(pattern, prevnick)

    if mat is not None:
        nick = f'{mat.group("nickname")}({id})'
    else:
        nick = f'{prevnick}({id})'

    await ctx.author.edit(nick=nick)
    await ctx.send(f'{prevnick}의 닉네임이 {nick}으로 변경되었습니다.')


@bot.command(aliases=['pos'], description="섬의 좌표와 위키 링크를 출력합니다.",
             help=f"{HELP_PREFIX}좌표`` ``섬 이름``")
async def 좌표(ctx, *, island):
    island = island.lower()

    iseng = re.match(r'[^ㄱ-ㅎㅏ-ㅣ가-힣]', island)
    iseng = True if iseng is not None else False

    if iseng:
        islands = list(engislandpos.keys())
    else:
        islands = list(islandspos.keys())

    substr = [x for x in islands if island in x.lower()]

    if (len(substr) == 0):
        clmat = difflib.get_close_matches(island, islands, cutoff=0.5)
        if (len(clmat) == 0):
            await ctx.send(f"``{island}`` 섬을 찾을 수 없습니다.")
            return None
        else:
            island = clmat[0]
    else:
        island = substr[0]
    await ctx.send(embed=make_pos_embed(island, iseng))


def make_pos_embed(name, iseng):
    if iseng:
        pr = engislandpos[name]
    else:
        pr = islandspos[name]
    urlname = pr[2].replace(" ", "_")
    WIKI_URL = f"https://seaofthieves.gamepedia.com/{urlname}"
    embed = discord.Embed(title=name, url=WIKI_URL)
    embed.add_field(name="좌표", value=pr[0], inline=True)
    embed.add_field(name="해역", value=pr[1], inline=True)
    embed.set_footer(text="섬 이름 클릭시 위키로 이동됨")
    return embed


lastchktime = None
lastserverstat = ""
RELOAD_TIME = datetime.timedelta(minutes=3)


def dt_to_str(date: datetime.datetime):
    return date.strftime("%Y-%m-%d %H:%M:%S")


@bot.command(description="서버 상태를 확인합니다. 쿨타임 3분",
             help=f"{HELP_PREFIX}서버``")
async def 서버(ctx):
    STATURL = "https://www.seaofthieves.com/ko/status"
    SERVER_IS_SICK = ":regional_indicator_x: 점검중이거나 서버가 터졌어요."
    SERVER_HEALTHY = ":white_check_mark: 서버가 정상이에요."
    SERVER_UNKNOWN = "서버 상태를 확인할 수 없어요."

    global lastserverstat, lastchktime
    msg = None

    if lastchktime is None:
        cachetime = RELOAD_TIME
    else:
        cachetime = datetime.datetime.now() - lastchktime

    if not cachetime < RELOAD_TIME:
        msg = await ctx.send("서버 상태 확인중...")
        async with aiohttp.ClientSession() as session:
            async with session.get(STATURL) as res:
                lastchktime = datetime.datetime.now()

                if res.status == 200:
                    soup = BeautifulSoup(await res.text(), 'html.parser')
                elif res.status != 200 or res is None:
                    lastserverstat = SERVER_UNKNOWN
                    await ctx.send(lastserverstat)
                    return

        srvinfo = soup.find("div", {"class": "service__info"})
        text = srvinfo.find("h2").contents[0]

        if text == "문제가 발생하여 서비스에 지장이 있습니다.":
            lastserverstat = SERVER_IS_SICK
        if text == "모든 서비스가 정상 운영되고 있습니다.":
            lastserverstat = SERVER_HEALTHY
        else:
            lastserverstat = SERVER_UNKNOWN

    lctstr = dt_to_str(lastchktime)

    if msg is not None:
        await msg.delete()

    await ctx.send(f"{lastserverstat} 확인 시간 : ``{lctstr}``")


bot.remove_command('help')


@bot.command(aliases=['help'], description='이 도움말을 출력합니다.')
async def 도움말(ctx, cmd):
    if cmd == 'help' or cmd == "도움말":
        await ctx.send(embed=make_help_embed())
    elif cmd in viscomsdict:
        await ctx.send(f"{viscomsdict[cmd].help}")
    else:
        await ctx.send(f'``{cmd}`` 명령어를 찾을 수 없습니다.')


def make_help_embed():
    coms = viscoms
    embed = discord.Embed(title='명령어 리스트')

    for c in coms:
        arr = [c.name]
        arr.extend(c.aliases)
        embed.add_field(name=', '.join(arr), value=c.description)

    embed.set_footer(text=f"'{CMD_PREFIX}help 명령어' 입력시 개별 명령어 도움말 출력")
    return embed
# endregion

# region logging

@bot.before_invoke
async def log_i(ctx: commands.Context):
    print(f"[info/{ctx.message.created_at}] {ctx.author} | {ctx.message.content}")


def log_e(ctx, error):
    now = dt_to_str(datetime.datetime.now())
    print(f"[error/{now}] | {ctx.message.content} | {error}")
#endregion

#region events


@bot.event
async def on_message(msg):
    allowchannel = [635398034469158914, 738655532709183551]
    # test channel, gall discord channel
    if msg.channel.id not in allowchannel:
        return
    await bot.process_commands(msg)


@bot.event
async def on_ready():
    await bot.change_presence(activity=discord.Game(name=f'{CMD_PREFIX}help'))
    print(f'봇 실행중 서버 : {len(bot.guilds)}개')


@bot.event
async def on_command_error(ctx, error):
    await log_i(ctx)
    if isinstance(error, commands.MissingRequiredArgument):
        if ctx.command.name == "도움말":
            await ctx.send(embed=make_help_embed())
        else:
            await ctx.send(f"사용법 : {ctx.command.help}")
    else:
        log_e(ctx, error)
        owner = f"<@!{bot.owner_id}>"
        await ctx.send(f"에러가 발생했습니다. {owner}")

#endregion

try:
    BOT_TOKEN = os.environ['BOT_TOKEN']
except KeyError:
    import config
    BOT_TOKEN = config.TOKEN

viscomsdict = {k: v for k, v in bot.all_commands.items() if not v.hidden}
viscoms = [x for x in bot.commands if not x.hidden]

bot.run(BOT_TOKEN)
