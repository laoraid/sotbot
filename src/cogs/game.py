import re
import difflib
import datetime
import aiohttp
import math

import discord
from discord.ext import commands
import openpyxl
from bs4 import BeautifulSoup
import tossi

from ..utils import dt_to_str, mkhelpstr
from .. import utils
from ..utils import converters
from ..config import CMD_PREFIX

RELOAD_TIME = datetime.timedelta(minutes=3)


class Game(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

        self.lastchktime = None
        self.lastserverstat = ""

        sheet = openpyxl.load_workbook(
            './src/islands.xlsx', data_only=True)['KOR']
        i = 2
        self.islandspos = {}
        self.engislandpos = {}
        while (True):
            islandname = sheet.cell(row=i, column=1).value

            if islandname is None:
                break
            pos = sheet.cell(row=i, column=2).value
            region = sheet.cell(row=i, column=3).value
            engname = sheet.cell(row=i, column=4).value
            chicken = bool(int(sheet.cell(row=i, column=5).value))
            pig = bool(int(sheet.cell(row=i, column=6).value))
            snake = bool(int(sheet.cell(row=i, column=7).value))

            v = [pos, region, engname, chicken, pig, snake]

            self.islandspos[islandname] = v
            self.engislandpos[engname] = v
            i += 1

    @commands.command(aliases=['pos'], description="섬의 좌표와 위키 링크를 출력합니다.",
                      usage=mkhelpstr("좌표", "섬 이름", aliases=["pos"]))
    async def 좌표(self, ctx, *, island):
        origin = island
        island = island.lower()

        iseng = re.match(r'[^ㄱ-ㅎㅏ-ㅣ가-힣]', island)
        iseng = True if iseng is not None else False

        if iseng:
            islands = list(self.engislandpos.keys())
        else:
            islands = list(self.islandspos.keys())

        substr = [x for x in islands if re.search(rf"\b{island}\b", x.lower())]
        if len(substr) == 0:
            substr = [x for x in islands if island in x.lower()]

        if (len(substr) == 0):
            clmat = difflib.get_close_matches(island, islands, cutoff=0.3)
            if (len(clmat) == 0):
                await ctx.send(f"``{island}`` 섬을 찾을 수 없습니다.")
                utils.log_v(ctx, f"{island} 검색 실패")
                return None
            else:
                island = clmat[0]
                ratio = difflib.SequenceMatcher(None, origin, island).ratio()
                utils.log_v(ctx, f"{origin} -> {island} (유사도 : {ratio:.2f})")
        else:
            island = substr[0]
            utils.log_v(ctx, f"{origin} -> {island} (부분 문자열)")
        await ctx.send(embed=self.make_pos_embed(island, iseng))

    def clr_region(self, region):
        RED = 0xd22b55
        GREEN = 0x05f445
        GRAY = 0x7c828e

        if region == "The Devil's Roar":
            return RED
        if region == "The Shores of Plenty":
            return GREEN
        if region == "The Wilds":
            return GRAY

        return utils.randcolor()

    def make_pos_embed(self, name, iseng):
        if iseng:
            pr = self.engislandpos[name]
        else:
            pr = self.islandspos[name]
        urlname = pr[2].replace(" ", "_")
        WIKI_URL = f"https://seaofthieves.gamepedia.com/{urlname}"
        embed = discord.Embed(title=name, url=WIKI_URL,
                              color=self.clr_region(pr[1]))
        embed.add_field(name="좌표", value=pr[0], inline=True)
        embed.add_field(
            name="동물", value=converters.convert_animal(pr), inline=True)
        embed.add_field(name="해역", value=pr[1], inline=False)
        embed.set_footer(text="섬 이름 클릭시 위키로 이동됨")
        return embed

    @commands.command(aliases=["server"], description="서버 상태를 확인합니다. 쿨타임 3분",
                      usage=mkhelpstr("서버", aliases=["server"]))
    async def 서버(self, ctx):
        STATURL = "https://www.seaofthieves.com/ko/status"
        SERVER_IS_SICK = ":regional_indicator_x: 점검중이거나 서버가 터졌어요."
        SERVER_HEALTHY = ":white_check_mark: 서버가 정상이에요."
        SERVER_UNKNOWN = "서버 상태를 확인할 수 없어요."

        msg = None

        if self.lastchktime is None:
            cachetime = RELOAD_TIME
        else:
            utc = datetime.datetime.utcnow()
            utc = utils.UTC.localize(utc)
            cachetime = utc - self.lastchktime

        if not cachetime < RELOAD_TIME:
            msg = await ctx.send("서버 상태 확인중...")
            async with aiohttp.ClientSession() as session:
                async with session.get(STATURL) as res:
                    self.lastchktime = datetime.datetime.utcnow()
                    self.lastchktime = utils.toKCT(self.lastchktime)

                    if res.status == 200:
                        soup = BeautifulSoup(await res.text(), 'html.parser')
                    elif res.status != 200 or res is None:
                        utils.log_v(ctx, "서버 상태 확인 불가")
                        self.lastserverstat = SERVER_UNKNOWN
                        await ctx.send(self.lastserverstat)
                        return None

            utils.log_v(ctx, "서버 상태 불러옴")
            srvinfo = soup.find("div", {"class": "service__info"})
            text = srvinfo.find("h2").contents[0]

            if text == "문제가 발생하여 서비스에 지장이 있습니다.":
                self.lastserverstat = SERVER_IS_SICK
            if text == "모든 서비스가 정상 운영되고 있습니다.":
                self.lastserverstat = SERVER_HEALTHY
            else:
                self.lastserverstat = SERVER_UNKNOWN
        else:
            utils.log_v(ctx, f"이전 상태 불러옴 {cachetime}")
        lctstr = dt_to_str(self.lastchktime)

        if msg is not None:
            await msg.delete()

        await ctx.send(f"{self.lastserverstat} 확인 시간 : ``{lctstr}``")

    def _calc_distance(self, pos, animals):
        def get_int(pos):
            x = ord(pos[0])
            y = int(pos[1:])
            return (x, y)

        def get_distance(x1, y1, x2, y2):
            dx = (x1 - x2) ** 2
            dy = (y1 - y2) ** 2

            return math.sqrt(dx + dy)

        def chk_animal(arr, animals):
            for animal in animals:
                if not arr[animal]:
                    return False
            return True

        stx, sty = get_int(pos)
        items = self.islandspos.items()
        islands = [(k, get_int(v[0].replace("-", "")))
                   for k, v in items if chk_animal(v, animals)]

        distemp = 100000
        numtemp = -1

        if len(islands) == 0:
            return None

        for i, arr in enumerate(islands):
            dis = get_distance(stx, sty, arr[1][0], arr[1][1])

            if dis < distemp:
                distemp = dis
                numtemp = i

        utils.log_v(None, f"{pos}에서 {islands[numtemp][0]}, 거리 : {distemp}")

        return islands[numtemp][0]

    @commands.command(description="현재 좌표에서 가장 가까운 동물이 있는 섬을 찾습니다.",
                      usage=mkhelpstr("동물", "현재 좌표", "동물1", "동물2..."))
    async def 동물(self, ctx, p: converters.Position,
                 n: commands.Greedy[converters.Animal]):
        if len(n) == 0:
            raise commands.BadArgument(
                f"동물 이름을 알 수 없습니다.\nEX) {CMD_PREFIX}동물 E21 닭 돼지")
        island = self._calc_distance(p, n)

        if island is None:
            await ctx.send("조건에 맞는 섬을 찾을 수 없습니다.")
            utils.log_v(ctx, "조건에 맞는 섬 검색 실패")
            return None

        animalstrs = [converters.decode_animal(x) for x in n]
        animalstrs = f"``{', '.join(animalstrs)}``"
        animalstrs = tossi.postfix(animalstrs, "이")

        await ctx.send(f"``{p}``에서 {animalstrs} 있는 가장 가까운 섬은...",
                       embed=self.make_pos_embed(island, False))

    @동물.error
    async def animal_error(self, ctx, error):
        if isinstance(error, commands.BadArgument):
            utils.log_e(ctx, error.args[0])
            await ctx.send(error.args[0])


def setup(bot):
    bot.add_cog(Game(bot))
